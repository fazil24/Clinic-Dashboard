import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import altair as alt
from datetime import datetime, timedelta

file_path = "C:/Users/DELL/Automation/Data.xlsx"

# Function to load existing Excel file or create a new one
def load_excel(filename):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        data = sheet.values
        columns = next(data)[0:]  # Get the first row as columns
        df = pd.DataFrame(data, columns=columns)
        
        # Ensure there are no duplicate columns
        df = df.loc[:, ~df.columns.duplicated()]

        # Reset the index to avoid any index-related errors
        df.reset_index(drop=True, inplace=True)
    except FileNotFoundError:
        df = pd.DataFrame(columns=["Name", "Age", "Phone Number", "Amount"])
    return df

# Function to save data to Excel
def save_to_excel(df, filename):
    #try:
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, header=False)
    # except FileNotFoundError:
    #     with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
    #         df.to_excel(writer, index=False, sheet_name='Sheet1')

#Generate date 

def generate_date_list(start_date, end_date):
    date_list = []
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)
    return date_list

# Function to display the dashboard
def display_dashboard(data):
    st.header("Dashboard")

    # KPI: Total Amount
    total_amount = data['Amount'].sum()
    st.metric("Total Amount", f"₹{total_amount:,.2f}")
    st.markdown('---')

    # Bar Chart: Amount by Name
    grouped_date = data.groupby('Date')['Amount'].sum().reset_index()
    bar_chart = alt.Chart(grouped_date).mark_bar().encode(
        x='Date:N',
        y='Amount:Q'
    ).properties(
        title='Amount by Date'
    )
    st.altair_chart(bar_chart, use_container_width=True)

    # Pie Chart: Distribution of Amount by Name
    grouped_df = data.groupby('Mode of Treatment')['Amount'].sum().reset_index()
    pie_chart = alt.Chart(grouped_df).mark_arc().encode(
        theta=alt.Theta(field="Amount", type="quantitative"),
        color=alt.Color(field="Amount", type="nominal"),
        tooltip=['Mode of Treatment', 'Amount']
    ).properties(
        title='Distribution of Amount by Mode of Treatment'
    )
    st.altair_chart(pie_chart, use_container_width=True)


# Streamlit interface
def main():
    st.title("User Information Form")

    #Date Setting
    start_date = datetime(2024, 8, 1)  # Start date for the dropdown
    end_date = datetime(2024, 8, 31)  # End date for the dropdown
    date_list = generate_date_list(start_date, end_date)
    date_options = [date.strftime('%d-%m-%Y') for date in date_list]  # Format dates as strings

    #Mode of Treatment setting
    mode_of_treatments_lists = ["Cinic","House Visit"]
    mode_of_treatment_options = [mode_of_treatment_list for mode_of_treatment_list in mode_of_treatments_lists]

    #Treatment setting
    list_of_treatment = ["IFT","WAX","PPL","ANOTHER"]
    treatment_options = [treatment_list for treatment_list in list_of_treatment]

    # Input fields
    date = st.selectbox("Date", options=date_options)
    name = st.text_input("Name")
    age = st.number_input("Age", min_value=1, max_value=120, step=1)
    phone = st.text_input("Phone Number")
    treatment = st.selectbox("Treatment", options=treatment_options)
    mode = st.selectbox("Mode of Treatment", options=mode_of_treatment_options)
    amount = st.number_input("Amount", min_value=0.0, step=0.01)

    if st.button("Submit"):
        # Load existing data
        data = load_excel(file_path)
        
        # Append new data
        new_data = pd.DataFrame([[date, name, age, phone, treatment, mode, amount]], 
                                columns=["Date", "Name", "Age", "Phone Number","Treatment","Mode of Treatment", "Amount"])
        #new_data = pd.concat([data, new_data], ignore_index=True)
        
        # Ensure the data has unique columns
        new_data = new_data.loc[:, ~new_data.columns.duplicated()]
        
        # Save to Excel
        save_to_excel(new_data, file_path)

        st.success("Data saved successfully!")

    # Load data and display the dashboard
    data = load_excel(file_path)
    if not data.empty:
        display_dashboard(data)

if __name__ == "__main__":
    main()